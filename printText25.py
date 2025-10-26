#


import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import time
from datetime import datetime

# 配置并启动 WebDriver
def start_driver():
    driver_path = ChromeDriverManager().install()
    service = Service(driver_path)
    options = webdriver.ChromeOptions()
    # options.add_argument("--headless")  # 无头模式
    options.add_argument("--disable-gpu")
    driver = webdriver.Chrome(service=service, options=options)
    return driver

# 提取第三行数字
def extract_target_number(cell_text):
    lines = cell_text.splitlines()
    if len(lines) >= 3:
        num_str = lines[2].replace(',', '')
        try:
            return int(num_str)
        except ValueError:
            return None
    return None

# 保存到 Excel 并生成编号列
def save_to_excel(collected_text, target_number, excel_file="output.xlsx"):
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    df_new = pd.DataFrame({
        "No": [None],
        "Time": [now],
        "Raw Text": [collected_text],
        "Target Number": [target_number]
    })

    try:
        df_old = pd.read_excel(excel_file, engine="openpyxl")
        df_all = pd.concat([df_old, df_new], ignore_index=True)
    except FileNotFoundError:
        df_all = df_new

    df_all["No"] = range(1, len(df_all) + 1)
    df_all.to_excel(excel_file, index=False, engine="openpyxl")
    print(f"数据已保存到 {excel_file}")

# 抓取推文数据
def fetch_tweet_data(tweet_url, excel_file='output.xlsx'):
    driver = start_driver()
    driver.get(tweet_url)
    time.sleep(5)

    collected_text = ""
    try:
        all_elements = driver.find_elements(By.XPATH, "//div | //span")
        for element in all_elements:
            text = element.text.strip()
            if "件の表示" in text:
                start_index = text.index("件の表示")
                collected_text = text[start_index:start_index + 30]
                break
    except Exception as e:
        print(f"抓取出错: {e}")
    finally:
        driver.quit()

    print("=== 抓取结果 ===")
    print(collected_text)
    print("================")

    target_number = extract_target_number(collected_text)
    save_to_excel(collected_text, target_number, excel_file)
    add_chart_to_excel(excel_file)

# 自动生成折线图
def add_chart_to_excel(file_name="output.xlsx"):
    wb = load_workbook(file_name)
    ws = wb.active

    # 查找列号
    header = {cell.value: cell.column for cell in ws[1]}
    if "Time" not in header or "Target Number" not in header:
        print("找不到所需列")
        return
    time_col = header["Time"]
    value_col = header["Target Number"]

    chart = LineChart()
    chart.title = "Target Number over Time"
    chart.x_axis.title = "Time"
    chart.y_axis.title = "Target Number"

    data = Reference(ws, min_col=value_col, min_row=1, max_row=ws.max_row)
    cats = Reference(ws, min_col=time_col, min_row=2, max_row=ws.max_row)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    ws.add_chart(chart, "F2")
    wb.save(file_name)
    print("图表已更新")

# 主函数
def main():
    tweet_url = "https://x.com/dramaDIVE_ytv/status/1981567801405161653"
    if not tweet_url:
        print("未输入 URL，程序退出")
        return
    fetch_tweet_data(tweet_url)

if __name__ == "__main__":
    main()
