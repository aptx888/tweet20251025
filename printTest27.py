#在本地电脑，每30分钟一次，电脑休眠也可以运行，电脑不关机


import pandas as pd

from openpyxl import load_workbook
from openpyxl.chart import ScatterChart, Reference, Series
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import time
from datetime import datetime

# 配置并启动 WebDriver
def start_driver():
    driver_path = ChromeDriverManager().install()
    service = Service(driver_path)
    options = webdriver.ChromeOptions()
    options.add_argument("--headless")  # 无头模式，可取消注释
    options.add_argument("--disable-gpu")
    driver = webdriver.Chrome(service=service, options=options)
    return driver

# 从单元格文本提取第三行数字
def extract_target_number(cell_text):
    lines = cell_text.splitlines()
    if len(lines) >= 3:
        num_str = lines[2].replace(',', '')
        try:
            return int(num_str)
        except ValueError:
            return None
    return None

# 抓取推文数据
def fetch_tweet_data(tweet_url):
    driver = start_driver()
    driver.get(tweet_url)
    time.sleep(5)

    collected_text = ""
    try:
        all_elements = driver.find_elements(By.XPATH, "//div | //span")
        for element in all_elements:
            text = element.text.strip()
            if "件の表示" in text:
                start_index = text.index("件の表示")
                collected_text = text[start_index:start_index + 30]
                break
    except Exception as e:
        print(f"抓取出错: {e}")
    finally:
        driver.quit()

    print("=== 抓取结果 ===")
    print(collected_text)
    print("================")
    return collected_text

# 保存数据到 Excel 并生成编号列
def save_to_excel(collected_text, excel_file="output.xlsx"):
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    target_number = extract_target_number(collected_text)

    new_row = {
        "No": None,
        "Time": now,
        "Raw Text": collected_text,
        "Target Number": target_number
    }

    try:
        df_old = pd.read_excel(excel_file, engine="openpyxl")
        df_all = pd.concat([df_old, pd.DataFrame([new_row])], ignore_index=True)
    except FileNotFoundError:
        df_all = pd.DataFrame([new_row])

    df_all["No"] = range(1, len(df_all) + 1)
    df_all.to_excel(excel_file, index=False, engine="openpyxl")
    print(f"数据已保存到 {excel_file}")

# 生成散点图
def add_scatter_chart_to_excel(file_name="output.xlsx"):
    wb = load_workbook(file_name)
    ws = wb.active

    # 获取列号
    header = {cell.value: cell.column for cell in ws[1]}
    if "Time" not in header or "Target Number" not in header:
        print("找不到所需列")
        return

    time_col = header["Time"]
    value_col = header["Target Number"]

    chart = ScatterChart()
    chart.title = "Target Number over Time"
    chart.x_axis.title = "Time"
    chart.y_axis.title = "Target Number"

    xvalues = Reference(ws, min_col=time_col, min_row=2, max_row=ws.max_row)
    yvalues = Reference(ws, min_col=value_col, min_row=2, max_row=ws.max_row)
    series = Series(yvalues, xvalues, title="Target Number")
    chart.series.append(series)

    ws.add_chart(chart, "F2")
    wb.save(file_name)
    print("散点图已更新")

# 主函数
#def main():
    #    tweet_url = "https://x.com/dramaDIVE_ytv/status/1981567801405161653"
    #    if not tweet_url:
    #   print("未输入网址，程序退出")
    #   return

    #text = fetch_tweet_data(tweet_url)
    #save_to_excel(text)
    #add_scatter_chart_to_excel()

# ... （所有 import, start_driver, extract_target_number, fetch_tweet_data 等函数保持不变） ...
# 主函数
def main():
    tweet_url = "https://x.com/dramaDIVE_ytv/status/1981567801405161653"
    if not tweet_url:
        print("未输入网址，程序退出")
        return

    print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] === 开始抓取 ===")
    text = fetch_tweet_data(tweet_url)
    save_to_excel(text)
    add_scatter_chart_to_excel()
    print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] === 抓取结束 ===")


if __name__ == "__main__":
    # 移除 while True 循环和 time.sleep(1800)
    # 任务计划程序会负责每 30 分钟调用一次这个脚本
    main()

# 确保将这个文件保存为一个名称，例如：`tweet_data_collector.py`
# 确保将 `output.xlsx` 和这个 `.py` 文件放在一个固定的文件夹中，例如 `C:\Scripts\TweetScraper\`








#if __name__ == "__main__":
#    while True:
 #       print("=== 开始抓取 ===")
  #      main()
   #     print("=== 等待30分钟 ===")
        #time.sleep(1800)  # 每30分钟执行一次
        #time.sleep(180)  # 每3分钟执行一次
    #    time.sleep(180)  # 每6分钟执行一次





